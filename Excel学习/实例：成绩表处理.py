import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('成绩表.xlsx')
sheet = wb['Sheet1']

for row in range(3, sheet.max_row+1):
    a = 0
    for columns in range(4, 7):
        cell = sheet.cell(row, columns)
        a += cell.value
    new_cell = sheet.cell(row, columns+1)
    new_cell.value = a
    new_cell2 = sheet.cell(row, columns+2)
    new_cell2.value = a/3

values = Reference(sheet,
                   min_row=3,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=columns+2)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'i2')

wb.save('成绩表2.xlsx')