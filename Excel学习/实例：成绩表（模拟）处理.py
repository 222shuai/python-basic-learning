import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('成绩表（模拟）.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
    a = 0
    for columns in range(2, 8):
        cell = sheet.cell(row, columns)
        a += cell.value
    new_cell = sheet.cell(row, columns + 1)
    new_cell.value = a
    new_cell2 = sheet.cell(row, columns + 2)
    new_cell2.value = a / 6

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=8,
                   max_col=9)
chat = BarChart()
chat.add_data(values)
sheet.add_chart(chat, 'j2')

wb.save('成绩表（模拟）2.xlsx')
