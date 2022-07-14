import openpyxl as xl  # Excel处理库
from openpyxl.chart import BarChart, Reference  # 条形图处理库


def process_workbook(filename):  # 定义了一个函数处理文件filename

    wb = xl.load_workbook(filename)  # 读取目标文件filename
    sheet = wb['Sheet1']  # 读取文件内目标表格（例：’Sheet1‘）

    for row in range(3, sheet.max_row + 1):  # ↓
        a = 0
        for columns in range(4, 7):
            cell = sheet.cell(row, columns)
            a += cell.value
        new_cell = sheet.cell(row, columns + 1)
        new_cell.value = a
        new_cell2 = sheet.cell(row, columns + 2)
        new_cell2.value = a / 3  # ↑ 表格数据处理更新

    values = Reference(sheet,  # 条形图数据收集
                       min_row=3,  # 数据最小行坐标
                       max_row=sheet.max_row,  # 数据最大行坐标
                       min_col=4,  # 数据最小列坐标
                       max_col=8)  # 数据最大列坐标

    chart = BarChart()  # 调用条形图生成函数
    chart.add_data(values)  # 条形图chart数据填充
    sheet.add_chart(chart, 'e2')  # 条形图安放位置(chart,坐标)(坐标：例‘e2’）

    wb.save(filename2)  # 保存文件
